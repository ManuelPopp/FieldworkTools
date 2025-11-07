import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

base_dir = "F:/FIELDWORK"
files_to_check = {
    "": [
        "{PLOT}.gpx", "{PLOT}_points.gpkg", "{PLOT}_boundary.gpkg",
        "{PLOT}_L2.kmz", "{PLOT}_M3M.kmz",
        "{PLOT}_report_L2.txt", "{PLOT}_report_M3M.txt"
    ],
    "Licor": [
        "Above/{PLOT}-A.txt",
        "Below/{PLOT}-B.txt",
        "{PLOT}_Processed_coords.xlsx"
    ],
    "DJITerra": [
        "GNDVI.tif", "LCI.tif", "NDRE.tif", "NDVI.tif", "OSAVI.tif",
        "result.tif", "result_Green.tif", "result_NIR.tif",
        "result_RedEdge.tif", "result_Red.tif",
        "cloud_merged.las", "dem.tif", "dom.tif", "dsm.tif",
        "dsm_m3m.tif"
    ]
}

folders = [
    f for f in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, f))
    ]

data = {
    folder: {
        os.path.join(sub, fname):
        os.path.exists(
            os.path.join(base_dir, folder, sub, fname.format(PLOT = folder))
            )
        for sub, fnames in files_to_check.items() for fname in fnames
    }
    for folder in folders
}

df = pd.DataFrame(data).T

df = df.rename(
    columns = {
        "{PLOT}.gpx": "GPX",
        "{PLOT}_points.gpkg": "Points",
        "{PLOT}_boundary.gpkg": "Boundary",
        "{PLOT}_L2.kmz": "L2 Mission",
        "{PLOT}_M3M.kmz": "M3M Mission",
        "{PLOT}_report_L2.txt": "L2 Report",
        "{PLOT}_report_M3M.txt": "M3M Report",
        "Licor\\Above/{PLOT}-A.txt": "Licor Above",
        "Licor\\Below/{PLOT}-B.txt": "Licor Below",
        "Licor\\{PLOT}_Processed_coords.xlsx": "Licor Processed",
        "DJITerra\\GNDVI.tif": "GNDVI",
        "DJITerra\\LCI.tif": "LCI",
        "DJITerra\\NDRE.tif": "NDRE",
        "DJITerra\\NDVI.tif": "NDVI",
        "DJITerra\\OSAVI.tif": "OSAVI",
        "DJITerra\\result.tif": "Result",
        "DJITerra\\result_Green.tif": "Green",
        "DJITerra\\result_NIR.tif": "NIR",
        "DJITerra\\result_RedEdge.tif": "RedEdge",
        "DJITerra\\result_Red.tif": "Red",
        "DJITerra\\cloud_merged.las": "Pointcloud",
        "DJITerra\\dem.tif": "DEM",
        "DJITerra\\dom.tif": "DOM",
        "DJITerra\\dsm.tif": "DSM",
        "DJITerra\\dsm_m3m.tif": "DSM M3M"
    })
excel_path = os.path.join(base_dir, "file_check.xlsx")
df.to_excel(excel_path)

wb = load_workbook(excel_path)
ws = wb.active
green = PatternFill(
    start_color = "C6EFCE", end_color = "C6EFCE", fill_type = "solid"
    )
red = PatternFill(
    start_color = "FFC7CE", end_color = "FFC7CE", fill_type = "solid"
    )

for r in range(2, ws.max_row + 1):
    for c in range(2, ws.max_column + 1):
        cell = ws.cell(row = r, column = c)
        if cell.value is True:
            cell.fill = green
        elif cell.value is False:
            cell.fill = red

wb.save(excel_path)